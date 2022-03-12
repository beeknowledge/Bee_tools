#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim:fenc=utf-8
#20220211 copyright BeeKnowledgeDesign

import os
from glob import glob
import openpyxl as excel
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import collections

classdList = {}
classDict ={}
dic2 ={}

wb = excel.Workbook()
ws = wb.active
ws.title = "Contents_table"
ws['A1']="ClassID"
ws['B1']="数量"
ws['C1']="ファイル総数"
ws['D1']="アノテーション総数"
num_totalClass = 0



global ln,classdata
classdata = []
ln  = 2 #　reset Column
dirName = "input" # Folder of annotation files for count


def replaceOb(fName,delLine):
    global ln,classdata
    with open(fName, "r") as input:
        for line in input:
            checkString = line.startswith(delLine)
            if checkString == True:
                mess =delLine.strip(" ")
                classdata.append(int(mess))
    input.close()

for file in glob(dirName + '/*.txt'):
    for n in range(23):
        replaceOb(file,str(n)+" ")


classdList = collections.Counter(classdata)
classDict = dict(classdList)
dic2 = sorted(classDict.items())

txtCounter = len(glob(dirName + '/*.txt'))
ws.cell( ln , 3 ).value = txtCounter # input to Excel <= number of files

for n in dic2:
    ws.cell( ln , 1 ).value = n[0] # input to Excel <= class number
    ws.cell( ln , 2 ).value = n[1]# input to Excel <= total
    ln += 1

cmax1 = ws.max_row
print(cmax1)
for i in range(2, cmax1+1):
    totalClass = ws['B' + str(i)].value
    num_totalClass = num_totalClass + totalClass

ws.cell( 2 , 4 ).value = num_totalClass # input to Excel <= number of num_totalClass

ln = 0
num_totalClass = 0
txtCounter = 0

side1 = Side(style='thin', color='000000')
border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)

wsPos = "A1:B" + str(cmax1)

for row in ws["C1:D2"]:
    for cell in row:
        cell.border = border_aro

for row in ws[wsPos]:
    for cell in row:
        cell.border = border_aro

fill = PatternFill(patternType='solid', fgColor='d3d3d3')

for row in ws["A1:D1"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="centerContinuous",shrinkToFit=True)
        ws[cell.coordinate].fill = fill

wb.save('投入データ数.xlsx')
